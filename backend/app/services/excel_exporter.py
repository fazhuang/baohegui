"""报告 Excel 导出服务。"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Mapping

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=16)
_BOLD_FONT = Font(bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _stringify(value: Any) -> Any:
    """将值转换为 Excel 单元格安全值。

    P0 安全：对所有以公式触发字符 (=, +, -, @) 开头的字符串值，
    在前面加单引号 ' 做前缀，防止 Excel Formula Injection (CWE-1236)。
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    s = str(value)
    # Prevent Excel / CSV formula injection (CWE-1236):
    # cells starting with =, +, -, @ can be interpreted as formulas.
    # Prefix with a single quote to force Excel to treat them as literal text.
    if s and s[0] in ('=', '+', '-', '@'):
        s = "'" + s
    return s


def _append_cell(ws, row: int, col: int, value: Any, *, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=_stringify(value))
    cell.alignment = _WRAP_TOP
    cell.border = _THIN_BORDER
    if bold:
        cell.font = _BOLD_FONT


def _write_table_header(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _set_column_widths(ws, widths: Mapping[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _risk_label(value: str) -> str:
    mapping = {"high": "高风险", "medium": "中风险", "low": "低风险"}
    return mapping.get(value, value or "")


def _rule_type_label(rule_type: str) -> str:
    mapping = {
        "chapter_required": "章节缺失",
        "keyword_required": "关键字缺失",
        "forbidden": "禁用词",
        "format_required": "格式要求",
    }
    return mapping.get(rule_type, rule_type or "")


def _llm_type_label(violation_type: str) -> str:
    mapping = {
        "exclusivity": "排他性",
        "bias": "倾向性",
        "hidden_barrier": "隐性壁垒",
        "ambiguity": "条款含糊",
        "high_risk": "质疑风险",
        "format_issue": "格式问题",
        "legal_risk": "法律风险",
        "procedural_issue": "程序问题",
    }
    return mapping.get(violation_type, violation_type or "")


def build_violation_rows(report_data: dict[str, Any]) -> list[dict[str, Any]]:
    """把报告 JSON 中的违规明细统一展平为 Excel 行。"""

    rows: list[dict[str, Any]] = []

    for item in report_data.get("rule_violations", []) or []:
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "source": "规则引擎",
                "rule_id": item.get("rule_id", ""),
                "rule_type": item.get("rule_type", ""),
                "risk_level": item.get("risk_level", ""),
                "category": _rule_type_label(item.get("rule_type", "")),
                "title": item.get("description", "") or item.get("rule_id", ""),
                "description": item.get("description", ""),
                "evidence_text": item.get("evidence_text") or item.get("text", ""),
                "suggestion": item.get("suggestion", ""),
                "law_ref": item.get("law_ref", ""),
                "confidence": item.get("confidence", item.get("template_confidence", "")),
            }
        )

    for item in report_data.get("llm_violations", []) or []:
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "source": "语义引擎",
                "rule_id": item.get("type", ""),
                "rule_type": item.get("type", ""),
                "risk_level": item.get("risk_level", ""),
                "category": _llm_type_label(item.get("type", "")),
                "title": item.get("reason", "") or item.get("type", ""),
                "description": item.get("reason", ""),
                "evidence_text": item.get("evidence", "") or item.get("text", ""),
                "suggestion": item.get("suggestion", ""),
                "law_ref": item.get("law_ref", ""),
                "confidence": item.get("confidence", ""),
                "validation_error": item.get("validation_error") or item.get("__validation_error__", ""),
                "requires_human_review": "是" if (item.get("requires_human_review") or item.get("__requires_human_review__")) else "否",
            }
        )

    return rows


def export_report_to_excel(report_data: dict[str, Any], violations: list[dict[str, Any]]) -> BytesIO:
    """将合规报告导出为 Excel（.xlsx）。"""

    wb = openpyxl.Workbook()

    # 概览
    ws_overview = wb.active
    ws_overview.title = "概览"
    ws_overview.sheet_view.showGridLines = False

    ws_overview["A1"] = "包合规审查报告"
    ws_overview["A1"].font = _TITLE_FONT
    ws_overview.merge_cells("A1:D1")

    info_rows = [
        ("文件名称", report_data.get("file_name", "")),
        ("审查时间", report_data.get("check_time", "")),
        ("总评分", report_data.get("total_score", 0)),
        ("章节完整性评分", report_data.get("section_score", 0)),
        ("关键字合规评分", report_data.get("keyword_score", 0)),
        ("禁用词检查评分", report_data.get("forbidden_score", 0)),
        ("语义审查评分", report_data.get("semantic_score", 0)),
        ("违规总数", report_data.get("total_violations", len(violations))),
        ("高风险数", report_data.get("high_risk_count", 0)),
        ("中风险数", report_data.get("medium_risk_count", 0)),
        ("低风险数", report_data.get("low_risk_count", 0)),
        ("LLM 模型", report_data.get("llm_model_used", "")),
        ("Token 消耗", report_data.get("llm_tokens_used", 0)),
        ("预估费用(元)", report_data.get("llm_cost_yuan", 0)),
        ("交通灯", report_data.get("traffic_light", "")),
        ("参数倾向性风险分", report_data.get("parameter_bias_score", "")),
        ("去重跨引擎数", report_data.get("dedup_cross_engine", "")),
        ("去重引擎内数", report_data.get("dedup_intra_engine", "")),
    ]

    for idx, (label, value) in enumerate(info_rows, start=3):
        _append_cell(ws_overview, idx, 1, label, bold=True)
        _append_cell(ws_overview, idx, 2, value)

    ws_overview.column_dimensions["A"].width = 20
    ws_overview.column_dimensions["B"].width = 40
    ws_overview.freeze_panes = "A3"

    # 违规明细
    ws_violations = wb.create_sheet("违规明细")
    ws_violations.sheet_view.showGridLines = False
    headers = [
        "来源",
        "规则ID",
        "规则类型",
        "风险等级",
        "分类",
        "标题",
        "描述",
        "证据原文",
        "整改建议",
        "法规引用",
        "信心度",
        "校验错误",
        "需人工复核",
    ]
    _write_table_header(ws_violations, headers)

    for row_idx, item in enumerate(violations, start=2):
        _append_cell(ws_violations, row_idx, 1, item.get("source", ""))
        _append_cell(ws_violations, row_idx, 2, item.get("rule_id", ""))
        _append_cell(ws_violations, row_idx, 3, item.get("rule_type", ""))
        _append_cell(ws_violations, row_idx, 4, _risk_label(str(item.get("risk_level", ""))))
        _append_cell(ws_violations, row_idx, 5, item.get("category", ""))
        _append_cell(ws_violations, row_idx, 6, item.get("title", ""))
        _append_cell(ws_violations, row_idx, 7, item.get("description", ""))
        _append_cell(ws_violations, row_idx, 8, item.get("evidence_text", ""))
        _append_cell(ws_violations, row_idx, 9, item.get("suggestion", ""))
        _append_cell(ws_violations, row_idx, 10, item.get("law_ref", ""))
        _append_cell(ws_violations, row_idx, 11, item.get("confidence", ""))
        _append_cell(ws_violations, row_idx, 12, item.get("validation_error", ""))
        _append_cell(ws_violations, row_idx, 13, item.get("requires_human_review", ""))

    ws_violations.freeze_panes = "A2"
    ws_violations.auto_filter.ref = f"A1:M{max(len(violations) + 1, 1)}"
    _set_column_widths(
        ws_violations,
        {
            "A": 12,
            "B": 14,
            "C": 16,
            "D": 12,
            "E": 14,
            "F": 30,
            "G": 42,
            "H": 36,
            "I": 32,
            "J": 20,
            "K": 10,
            "L": 30,
            "M": 14,
        },
    )

    # 引擎诊断
    ws_diag = wb.create_sheet("引擎诊断")
    ws_diag.sheet_view.showGridLines = False
    diag = report_data.get("_diagnostics", {}) or {}

    diag_rows: list[tuple[str, Any]] = [
        ("解析状态", "—"),
        ("解析质量", report_data.get("parse_quality", "ok")),
        ("章节数", diag.get("parser", {}).get("sections_found", 0)),
        ("页数", diag.get("parser", {}).get("page_count", 0)),
        ("", ""),
        ("路由判定", "—"),
        ("交通灯", diag.get("routing", {}).get("traffic_light", "")),
        ("跳过LLM", diag.get("routing", {}).get("skip_llm", False)),
        ("", ""),
        ("LLM审查统计", "—"),
        ("Provider", diag.get("llm_engine", {}).get("provider", "")),
        ("模型", diag.get("llm_engine", {}).get("model", "")),
        ("Token消耗", diag.get("llm_engine", {}).get("tokens_used", 0)),
        ("费用(元)", diag.get("llm_engine", {}).get("cost_yuan", 0)),
        ("", ""),
        ("耗时", "—"),
        ("总耗时(s)", diag.get("timing", {}).get("total_seconds", 0)),
        ("规则引擎(ms)", diag.get("timing", {}).get("rules_ms", 0)),
        ("LLM(ms)", diag.get("timing", {}).get("llm_ms", 0)),
    ]

    section_labels = {"解析状态", "路由判定", "LLM审查统计", "耗时"}
    for row_idx, (label, value) in enumerate(diag_rows, start=1):
        _append_cell(ws_diag, row_idx, 1, label, bold=label in section_labels)
        _append_cell(ws_diag, row_idx, 2, value)

    ws_diag.column_dimensions["A"].width = 18
    ws_diag.column_dimensions["B"].width = 24
    ws_diag.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
